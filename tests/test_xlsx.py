import os.path
from openpyxl import load_workbook
from conftest import RESOURSES_PATH

# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_reading_xlsx():
    workbook = load_workbook(os.path.join(RESOURSES_PATH, "file_example_XLSX_50.xlsx"))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    headers = []
    for i in range(8):
        headers.append(sheet.cell(row=1, column=i + 1).value)
        i += 1

    assert headers == [
        0,
        "First Name",
        "Last Name",
        "Gender",
        "Country",
        "Age",
        "Date",
        "Id",
    ]
